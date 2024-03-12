import os
from rapidfuzz import fuzz

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook



def get_sheet(path_file: str, list_name: str):
    workbook = load_workbook(path_file)
    if not list_name or list_name == '':
        return workbook.active, workbook
    try:
        return workbook[list_name], workbook
    except Exception as e:
        raise ValueError(e)


def get_values_by_column(ws: Workbook, col1: str, col2: str):
    try:
        values_1 = [cell.value for cell in ws[col1]]
        values_2 = [cell.value for cell in ws[col2]]

        return values_1, values_2
    except Exception as e:
        raise ValueError(e)


def get_matches_percent(str1: str, str2: str):
     result = fuzz.QRatio(str1, str2)
     return result


def compare(values_1, values_2):
    result = []
    for i in range(1,len(values_1)):
        str_1 = values_1[i]
        str_2 = values_2[i]
        percent = get_matches_percent(str_1, str_2)
        result.append((i, percent))
    return result


def write_resutl_in_excel(ws, wb, result, column):
    for item in result:
        a = ws.cell(row=item[0]+1, column=column, value=int(item[1]))
    wb.save("new.xlsx")



def main():

    path = "data.xlsx" # путь до файла
    name_list  = "Лист1" # название листа
    col1 = "C" # буква столбца сравниваемой колонки
    col2 = "D" # буква столбца сравниваемой колонки
    insert_column_number = 6 # номер столбца, куда будут записаны данные

    print("Загрука данных...")
    ws, wb = get_sheet(path, name_list)
    values_1, values_2 = get_values_by_column(ws, col1, col2)

    print("Сравнение ячеeк...")
    result = compare(values_1, values_2)

    print("Запись данных...")
    write_resutl_in_excel(ws, wb, result, insert_column_number)

    print("Обработка данных завершена.")

if __name__ == "__main__":
    main()
